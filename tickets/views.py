from django.shortcuts import *
from .models import *
from .forms import *
from django.contrib.auth.decorators import *
from django.contrib.auth import *
from datetime import *
from django.utils import timezone
from django.utils.timezone import *
from django.contrib import messages
from django.http import *
from django.utils.http import urlsafe_base64_encode
from django.dispatch import receiver
from django.db.models.signals import post_save
from django.views.decorators.csrf import csrf_exempt
from django.core.mail import send_mail
from django.db.models import *
from django.template.loader import render_to_string
from django.utils.html import strip_tags
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill
import os
import uuid
from django.conf import settings
from django.db.models.functions import ExtractMonth
from django.contrib.auth.models import Group
from django.db.models import *
import datetime
import xml.etree.ElementTree as ET
from decimal import Decimal
from django.contrib.auth.tokens import default_token_generator
from django.contrib.sites.shortcuts import get_current_site
from django.utils.encoding import force_bytes
from django.contrib.auth.models import User
import re
from datetime import *
import pandas as pd
import csv


    
@login_required
def home(request):
    # Exemplo de mensagem de boas-vindas
    analyst_tickets = Ticket.objects.filter(status='New').values('assigned_to').annotate(total=Count('id'))
    date_tickets = Ticket.objects.filter(status='New').values('created_at__date').annotate(total=Count('id'))
    companies = Company.objects.all()
    users = CustomUser.objects.all()
        # Calcular contagem de tickets do mês atual
    today = datetime.today()

    current_year = datetime.now().year
    monthly_tickets = Ticket.objects.filter(created_at__year=current_year).exclude(status='Cancelled') \
        .values('created_at__month').annotate(count=Count('id'))

    # Calcular contagem de tickets do ano atual
    start_of_year = today.replace(month=1, day=1, hour=0, minute=0, second=0, microsecond=0)
    annual_tickets = Ticket.objects.filter(status='New', created_at__gte=start_of_year).count()

    # Calcular contagem de tickets em aberto
    open_tickets = Ticket.objects.exclude(status='Closed').count()

        # Calcular contagem de tickets fechados
    closed_tickets = Ticket.objects.filter(status='Closed').count()

    context = {
        'companies': companies,
        'users': users,
        'analyst_tickets': analyst_tickets,
        'date_tickets': date_tickets,
        'monthly_tickets': monthly_tickets,
        'annual_tickets': annual_tickets,
        'open_tickets': open_tickets,
        'closed_tickets': closed_tickets,
    }
    # Renderiza o template 'home.html' passando a mensagem de boas-vindas como contexto
    return render(request, 'geral/home.html', context)

@login_required
def configuration_view(request):
    # Verifica se o usuário autenticado é um superusuário
    if not request.user.is_superuser:
        # Se não for, redireciona para a página inicial ou exibe uma mensagem de erro.
        # Por exemplo, redirecionar para a página de tickets (caso exista):
        return redirect('ticket_list')  # Certifique-se de ajustar o nome da URL de acordo com a sua aplicação

    # Se o usuário for um superusuário, renderize a página de configuração
    return render(request, 'geral/configuration.html')

def login_view(request):
    if request.method == 'POST':
        email = request.POST.get('email')
        password = request.POST.get('password')
        
        # Autenticar usando o email em vez do username
        user = authenticate(request, email=email, password=password)
        
        if user is not None:
            login(request, user)
            return redirect('home')
        else:
            messages.error(request, 'Email ou senha inválidos.')
    
    return render(request, 'user/login.html')

def user_list(request):
    users = CustomUser.objects.all()
    return render(request, 'user/user_list.html', {'users': users})

@login_required
def user_create(request):
    if request.method == 'POST':
        form = UserForm(request.POST)
        if form.is_valid():
            user = form.save(commit=False)
            
            # Extrai o primeiro nome e o último nome do formulário
            first_name = form.cleaned_data['first_name']
            last_name = form.cleaned_data['last_name']
            
            # Cria o username no formato "primeiro_nome.ultimo_nome"
            username = f"{first_name}.{last_name}"
            user.username = username
            
            password = form.cleaned_data['password']
            user.set_password(password)
            user.save()
            return redirect('user_list')
        else:
            print(form.errors)
            # Aqui vamos incluir a mensagem de erro no contexto
            context = {
                'form': form,
                'error_message': 'Erro ao criar usuário. Verifique os campos e tente novamente.',
            }
            return render(request, 'user/user_form.html', context)
    else:
        form = UserForm()

    context = {
        'form': form,
    }
    return render(request, 'user/user_form.html', context)

def auth_cpf(request):
    if request.method == 'POST':
        form = AuthCpfForm(request.POST)
        if form.is_valid():
            form.save()
           
            return redirect('home') 
    else:
        form = AuthCpfForm()

    context = {
        'form': form,
    }

    return render(request, 'user/auth_cpf.html', context)



def driver_registration(request):
    banks = Bank.objects.all()
    error_message = None
    
    if request.method == 'POST':
        form = DriverRegistrationForm(request.POST)
        
        if form.is_valid():
            first_name = form.cleaned_data['first_name']
            last_name = form.cleaned_data['last_name']
            email = form.cleaned_data['email']
            cpf = form.cleaned_data['cpf']
            cnpj = form.cleaned_data['cnpj']
            password = form.cleaned_data['password']
            account_digit = form.cleaned_data['account_digit']
            account_number = form.cleaned_data['account_number']
            account_type = form.cleaned_data['account_type']
            branch = form.cleaned_data['branch']
            bank = form.cleaned_data['bank_id']

            # Verifique se o CPF ou CNPJ está na tabela AuthCpf
            if not AuthCpf.objects.filter(cpf=cpf).exists() and not AuthCpf.objects.filter(cnpj=cnpj).exists():
                error_message = "CPF ou CNPJ não autorizado"
            else:
                # Crie o username no formato "primeiro_nome.ultimo_nome"
                username = f"{first_name}.{last_name}"

                # Crie um novo CustomUser
                custom_user = CustomUser.objects.create(
                    username=username,
                    email=email,
                    first_name=first_name,
                    last_name=last_name,
                    is_driver=True,
                    is_staff=True,
                    account_digit=account_digit,
                    account_number=account_number,
                    account_type=account_type,
                    bank=bank,
                    cnpj=cnpj,
                    cpf=cpf,
                    branch=branch,
                )
                
                # Configure a senha do CustomUser
                custom_user.set_password(password)
                custom_user.save()
                
                # Crie uma instância do modelo Drivers
                driver_instance = Drivers(
                    driver=custom_user,
                    cpf=cpf,
                    cnpj=cnpj,
                    bank=bank,  # Atribua a instância do banco selecionado
                    branch=branch,
                    account_type=account_type,
                    account_number=account_number,
                    account_digit=account_digit,
                )
                driver_instance.save()
                
                return redirect('home')  # Redirecione para a lista de drivers ou outra página desejada
        else:
            print(form.errors)
    else:
        form = DriverRegistrationForm()

    context = {
        'form': form,
        'banks': banks,
        'error_message': error_message  # Passe a mensagem de erro para o contexto
    }
    return render(request, 'user/register_driver.html', context)


def user_edit(request, pk):
    user = get_object_or_404(CustomUser, pk=pk)
    
    if request.method == 'POST':
        form = UserForm(request.POST, instance=user)
        
        # Verifique se a senha está vazia
        if 'password' in request.POST and not request.POST['password']:
            form.fields['password'].required = False  # Torna a senha opcional
            form.fields['password'].widget.attrs['disabled'] = True  # Desativa o campo senha

        if form.is_valid():
            # Salve apenas se a senha não estiver vazia
            if 'password' in request.POST and not request.POST['password']:
                user.password = form.instance.password
            else:
                user.set_password(form.cleaned_data['password'])

            form.save()
            return redirect('user_list')
    else:
        form = UserForm(instance=user)

    return render(request, 'user/user_edit.html', {'form': form})


def user_delete(request, pk):
    user = get_object_or_404(CustomUser, pk=pk)
    if request.method == 'POST':
        user.delete()
        return redirect('user_list')
    return render(request, 'user/user_confirm_delete.html', {'user': user})


def password_reset_done(request):
    return render(request, 'registration/password_reset_done.html')

def password_reset_complete(request):
    return render(request, 'registration/password_reset_complete.html')



def reset_password(request):
    if request.method == "POST":
        form = SetPasswordForm(request.POST)
        if form.is_valid():
            email = form.cleaned_data['email']
            user = User.objects.filter(email=email).first()
            if user:
                uid = urlsafe_base64_encode(force_bytes(user.pk))
                token = default_token_generator.make_token(user)
                current_site = get_current_site(request)
                mail_subject = 'Reset your password'
                message = render_to_string('registration/password_reset_email.html', {
                    'user': user,
                    'domain': current_site.domain,
                    'uid': uid,
                    'token': token,
                    'protocol': 'https' if request.is_secure() else 'http',
                })
                send_mail(mail_subject, message, 'your_email@gmail.com', [email])
            return redirect('password_reset_done')
    else:
        form = SetPasswordForm()
    return render(request, 'registration/password_reset_form.html', {'form': form})


def ticket_list(request):
    tickets = Ticket.objects.annotate(comments_count=Count('comments'))
    return render(request, 'tickets/ticket_list.html', {'tickets': tickets})

def ticket_detail(request, ticket_id):
    ticket = get_object_or_404(Ticket, id=ticket_id)
    users = CustomUser.objects.all()
    other_users = users.exclude(pk=request.user.id)

    if request.method == 'POST':
        form = CommentForm(request.POST)
        change_assigned_form = ChangeAssignedToForm(request.POST)

        if form.is_valid() and change_assigned_form.is_valid():
            comment = form.cleaned_data['comment']
            ticket.comments.create(author=request.user, body=comment, ticket=ticket)

            selected_users = request.POST.getlist('to')
            for user_id in selected_users:
                user = CustomUser.objects.get(pk=user_id)
                if user != request.user:
                    send_mail_to_user(user, ticket, comment)

            return redirect('ticket_detail', ticket_id=ticket_id)
    else:
        form = CommentForm()
        change_assigned_form = ChangeAssignedToForm()

    context = {
        'ticket': ticket,
        'users': users,
        'other_users': other_users,
        'form': form,
        'change_assigned_form': change_assigned_form,
    }
    return render(request, 'tickets/ticket_detail.html', context)

def send_mail_to_user(user, ticket, comment):
    subject = f"Notificação sobre o ticket {ticket.id}"
    from_email = 'seu_email@exemplo.com'  # Substitua pelo seu e-mail
    recipient_list = [user.email]
    
    # Carregar o template de e-mail e substituir os campos dinâmicos pelo valor correto
    context = {
        'user': user,  # Aqui utilizamos o objeto CustomUser
        'ticket': ticket,
        'comment': comment,
    }
    html_message = render_to_string('tickets/email/comment_email.html', context)
    plain_message = strip_tags(html_message)
    
    send_mail(subject, plain_message, from_email, recipient_list, html_message=html_message)


def sla_calculation(created_at, priority):
    if priority == 'Low':
        prazo = 48
    elif priority == 'Medium':
        prazo = 24
    elif priority == 'High':
        prazo = 4
    elif priority == 'Urgent':
        prazo = 2
    elif priority == 'Scheduled':
        prazo = 72
    else:
        raise ValueError('Invalid Priority.')

    # Adiciona o prazo em horas à data de criação
    sla_date = created_at + timedelta(hours=prazo)

    # Verifica se a data limite cai em um sábado ou domingo
    while sla_date.weekday() >= 5:
        # Se cair em um fim de semana, ajusta para o início da próxima semana (segunda-feira)
        sla_date += timedelta(days=(7 - sla_date.weekday()))

    # Define o horário de início e fim do expediente
    start_time = time(9, 0, 0)
    end_time = time(18, 0, 0)

    # Verifica se a data limite está fora do horário de trabalho
    if sla_date.time() < start_time:
        sla_date = datetime.combine(sla_date.date(), start_time)
    elif sla_date.time() > end_time:
        sla_date = datetime.combine(sla_date.date() + timedelta(days=1), start_time)

    return sla_date


#Função original de criação de ticket (individual)
@login_required
def ticket_create(request):
    if request.method == 'POST':
        form = TicketForm(request.POST, request.FILES)
        if form.is_valid():
            ticket = form.save(commit=False)
            ticket.created_by = request.user

            # Verifica se o campo company foi preenchido no formulário
            if 'company' in form.cleaned_data and form.cleaned_data['company']:
                ticket.company = form.cleaned_data['company']

            # Calcular data de criação do ticket
            ticket.created_at = make_aware(datetime.now())

            # Calcular data limite (SLA) com base na prioridade selecionada
            ticket.sla_date = sla_calculation(ticket.created_at, ticket.priority)

            ticket.save()
            print("Ticket created successfully!")
            return redirect('ticket_detail', ticket_id=ticket.id)
        else:
            print("Form is not valid:", form.errors)
    else:
        form = TicketForm()

    companies = Company.objects.all()
    users = CustomUser.objects.all()

    context = {
        'form': form,
        'companies': companies,
        'users': users,
    }

    return render(request, 'tickets/ticket_form.html', context)

#Função de sinal, observa o DeliveryCalendar, quando é criado gera os tickets abaixo
@receiver(post_save, sender=DeliveryCalendar)
def create_tickets(sender, instance, created, **kwargs):
    if created:
        # Mapeamento das atividades para os títulos
        activity_titles = {
            'Changes': 'Payroll Changes',
            'Payroll Preview': 'Payroll Preview',
            'Approval': 'Payroll Approval',
            'Net Salaries': 'Net Salaries',
            'Taxes': 'Taxes',
            'Accounting': 'Accounting and Provisions',
            'Next Month Vacations': 'Next Month Vacations',
            'eSocial Delivery': 'eSocial Delivery',
        }

        cam = instance.cam
        analyst = instance.analyst
        ticket_sla_date = instance.month_date

        if instance.pay_period == 'Advance':
            # Cria as atividades para pay_period igual a 'Advance'
            activity_titles = {
                'Changes': 'Payroll Changes',
                'Payroll Preview': 'Payroll Preview',
                'Approval': 'Payroll Approval',
                'Net Salaries': 'Net Salaries',
            }
        
        for activity, title in activity_titles.items():
            activity_date_field = f'{activity.replace(" ", "_").lower()}_date'
            sla_date = getattr(instance, activity_date_field, ticket_sla_date)
            
            Ticket.objects.create(
                title=title,
                description=f'This is an automatic ticket for {activity} activity',
                priority='Medium',
                assigned_to=analyst,
                created_by=cam,
                company=instance.client,
                department='Operation',
                type=activity,
                sla_date=sla_date,
            )

#nessa função é para criar um ticket automaticamente quando um ticket do tipo "termination" é criado e solicitar ao usuario que tire os extratos de FGTS
@receiver(post_save, sender=Ticket)
def create_tickets(sender, instance, created, **kwargs):
    if created and instance.type == 'Termination':
    
        # Mapeamento das atividades para os títulos
        activity_titles = {
            'FGTS': f'FGTS Statement Request from Ticket #{instance.id}',
        }
        cam = instance.created_by
        analyst = instance.assigned_to
        ticket_sla_date = instance.sla_date - timedelta(days=1)  # Subtrair um dia do sla_date
        
        for activity, title in activity_titles.items():
            Ticket.objects.create(
                title=title,
                description=f'This is an automatic ticket for {activity} activity. Please access https://conectividadesocialv2.caixa.gov.br/sicns/ and request for detailed FGTS balance statement for termination calculation.',
                priority='Medium',
                assigned_to=analyst,
                created_by=cam,
                company=instance.company,
                department='Operation',
                type=activity,
                sla_date=getattr(instance, f'{activity}_date', ticket_sla_date),
            )

#aqui a função deve ser cria um ticket a partir de alteração de status do ticket de Vacation para 'closed', criar um ticket para o CAM enviar os entregaveis ao cliente.
@receiver(post_save, sender=Ticket)
def create_tickets(sender, instance, created, **kwargs):
    if not created and instance.type == 'Vacation' and instance.status == 'Closed':
        # Mapeamento das atividades para os títulos
        activity_titles = {
            'Vacation': 'Vacation Activity Completed',
        }
        cam = instance.assigned_to
        analyst = instance.created_by
        ticket_sla_date = instance.closed_at  # Usar a data de conclusão do ticket anterior
        
        for activity, title in activity_titles.items():
            Ticket.objects.create(
                title=title,
                description='The Vacation activity has been completed. Please send the deliverables to the client.',
                priority='Urgent',
                assigned_to=cam,
                created_by=analyst,
                company=instance.company,
                department='Services',
                type=activity,
                sla_date=ticket_sla_date,
            )

def ticket_list_all_closed(request):
    tickets = Ticket.objects.filter(status='Closed').annotate(comments_count=Count('comments'))

    return render(request, 'tickets/ticket_list_closed.html', {'tickets': tickets})

def ticket_list_all_open(request):
    tickets = Ticket.objects.exclude(status='Closed').annotate(comments_count=Count('comments'))
    return render(request, 'tickets/ticket_list_all_open.html', {'tickets': tickets})

def ticket_list_my_open(request):
    user = request.user
    tickets = Ticket.objects.filter(assigned_to=user).exclude(status='Closed').annotate(comments_count=Count('comments'))

    return render(request, 'tickets/ticket_list_my_open.html', {'tickets': tickets})

def ticket_list_my_closed(request):
    # Filtra os tickets em aberto do usuário logado
    user = request.user
    tickets = Ticket.objects.filter(assigned_to=user).filter(status='Closed').annotate(comments_count=Count('comments'))

    return render(request, 'tickets/ticket_list_my_closed.html', {'tickets': tickets})

def open_tickets_report_xlsx(request):
    open_tickets = Ticket.objects.exclude(status='Closed')

    # Dicionário de cores para cada valor de prioridade
    priority_colors = {
        'low': '008000',    # Verde
        'medium': 'FFFF00', # Amarelo
        'high': 'FFA500',   # Laranja
        'urgent': 'FF0000', # Vermelho
        'scheduled': '0000FF', # Azul
    }


    # Cria um novo arquivo do Excel
    workbook = Workbook()
    worksheet = workbook.active

    # Define os títulos das colunas
    worksheet['A1'] = 'ID'
    worksheet['B1'] = 'Company'
    worksheet['C1'] = 'Partner'
    worksheet['D1'] = 'Title'
    worksheet['E1'] = 'Requester'
    worksheet['F1'] = 'Owner'
    worksheet['G1'] = 'Created Date'
    worksheet['H1'] = 'SLA Date'
    worksheet['I1'] = 'Priority'
    worksheet['J1'] = 'Link'

    # Define o alinhamento do texto como centralizado
    for column in worksheet.columns:
        for cell in column:
            cell.alignment = Alignment(horizontal='center')

    # Preenche os dados dos tickets abertos no relatório
    row_index = 2
    for ticket in open_tickets:
        worksheet.cell(row=row_index, column=1, value=ticket.id)
        worksheet.cell(row=row_index, column=2, value=ticket.company.abreviation if ticket.company and hasattr(ticket.company, 'name') else '')
        worksheet.cell(row=row_index, column=3, value=ticket.company.partner)
        worksheet.cell(row=row_index, column=4, value=ticket.title)
        worksheet.cell(row=row_index, column=5, value=ticket.created_by.first_name)
        worksheet.cell(row=row_index, column=6, value=ticket.assigned_to.first_name)
        worksheet.cell(row=row_index, column=7, value=ticket.created_at.replace(tzinfo=None))
        worksheet.cell(row=row_index, column=8, value=ticket.sla_date.replace(tzinfo=None) if ticket.sla_date else None)
        worksheet.cell(row=row_index, column=9, value=ticket.priority)
        worksheet.cell(row=row_index, column=10, value=ticket.link)

        # Aplica a formatação condicional na coluna de prioridade
        priority = ticket.priority.lower()
        color = priority_colors.get(priority, 'FFFFFF')  # Branco se a prioridade não estiver mapeada
        cell = worksheet.cell(row=row_index, column=9)
        fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
        cell.fill = fill

        # Verifica se a data SLA é igual à data de hoje e aplica a cor verde
        sla_date = ticket.sla_date.date() if ticket.sla_date else None
        if sla_date == date.today():
            cell.fill = PatternFill(start_color='008000', end_color='008000', fill_type='solid')  # Cor verde

        row_index += 1

    unique_id = str(uuid.uuid4())[:5]  # Obtém os primeiros 5 caracteres do UUID
    
    # Define o nome do arquivo do relatório com o identificador único
    report_filename = f'Open_Tickets_{unique_id}.xlsx'

    # Define o caminho completo onde o arquivo será salvo
    report_path = os.path.join(settings.MEDIA_ROOT, 'reports', report_filename)

    # Salva o arquivo do Excel
    workbook.save(report_path)

    # Envia o arquivo como resposta para o navegador
    with open(report_path, 'rb') as file:
        response = HttpResponse(file.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=' + report_filename
        return response

@csrf_exempt
def change_ticket_status(request):
    if request.method == 'POST':
        ticket_id = request.POST.get('ticket_id')
        new_status = request.POST.get('new_status')

        try:
            ticket = Ticket.objects.get(pk=ticket_id)

            if ticket.status == 'Closed' and new_status != 'Closed':
                # If the ticket is closed, only superusers can change it to another status
                if not request.user.is_superuser:
                    return JsonResponse({'success': False, 'error': 'Only superusers can change the status of a closed ticket.'})
                ticket.closed_by = None  # Remove the user who closed the ticket

            if ticket.status != 'Closed' and new_status == 'Closed':
                ticket.closed_by = request.user  # Save the instance of the logged-in user

            ticket.status = new_status
            ticket.save()
            return JsonResponse({'success': True})
        except Ticket.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Ticket not found'})
    return JsonResponse({'success': False, 'error': 'Invalid request method'})

def change_ticket_assigned_to(request, ticket_id):
    ticket = get_object_or_404(Ticket, id=ticket_id)

    if request.method == 'POST':
        form = ChangeAssignedToForm(request.POST)
        if form.is_valid():
            assigned_to = form.cleaned_data['assigned_to']
            ticket.assigned_to = assigned_to
            ticket.save()
            return redirect('ticket_detail', ticket_id=ticket.id)
    else:
        form = ChangeAssignedToForm()

    return render(request, 'tickets/change_assigned_to.html', {'form': form})

def ticket_close(request, ticket_id):
    try:
        ticket = Ticket.objects.get(id=ticket_id)
    except Ticket.DoesNotExist:
        # Lógica para tratamento de ticket não encontrado (opcional)
        return redirect('ticket_list')  # Redireciona de volta para a lista de tickets

    # Verifica se o ticket já está fechado
    if ticket.status == 'closed':
        # Lógica para tratamento de ticket já fechado (opcional)
        return redirect('ticket_list')  # Redireciona de volta para a lista de tickets

    # Atualiza os campos de fechamento do ticket
    ticket.closed_at = now()
    ticket.closed_by = request.user
    ticket.status = 'Closed'
    ticket.save()

    # Redireciona de volta para a lista de tickets após o fechamento
    return redirect('ticket_list_all_open')

def delete_tickets_by_id(ticket_ids):
    ticket_ids = ticket_ids.split(',')

    deleted_tickets = []
    failed_tickets = []

    for ticket_id in ticket_ids:
        if '-' in ticket_id:
            start, end = map(int, ticket_id.split('-'))
            for id in range(start, end + 1):
                try:
                    ticket = Ticket.objects.get(pk=id)
                    ticket.delete()
                    deleted_tickets.append(id)
                except Ticket.DoesNotExist:
                    failed_tickets.append(id)
        else:
            try:
                ticket = Ticket.objects.get(pk=int(ticket_id))
                ticket.delete()
                deleted_tickets.append(int(ticket_id))
            except Ticket.DoesNotExist:
                failed_tickets.append(int(ticket_id))

    return deleted_tickets, failed_tickets

def delete_tickets(request):
    if request.method == 'POST':
        ticket_ids = request.POST.get('ticket_ids', '')
        deleted_tickets, _ = delete_tickets_by_id(ticket_ids)

        return redirect('ticket_list')
    else:
        return render(request, 'tickets/delete_tickets_form.html')

def company_list(request):

    companies = Company.objects.all()
    return render(request, 'company/company_list.html', {'companies': companies})

@login_required
def company_create(request):
    if request.method == 'POST':
        form = CompanyForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('company_list')
    else:
        form = CompanyForm()

    cam_users = CustomUser.objects.filter(is_cam=True)
    analyst_users = CustomUser.objects.filter(is_analyst=True)

    context = {
        'cam_users': cam_users,
        'analyst_users': analyst_users,
        'form': form
    }

    return render(request, 'company/company_form.html', context)

def company_detail(request, company_id):
    company = Company.objects.get(pk=company_id)
    deliverable = Deliverable.objects.all()

    if request.method == 'POST':
        form = DeliverableForm(request.POST)
        if form.is_valid():
            deliverable = form.save(commit=False)
            deliverable.company = company
            deliverable.save()
            return redirect('company_detail', company_id=company_id)
    else:
        form = DeliverableForm()

    return render(request, 'company/company_detail.html', {'company': company, 'form': form, 'deliverables': deliverable})


def company_edit(request, company_id):
    # Obtém a empresa com base no ID ou retorna um erro 404 se não existir
    company = get_object_or_404(Company, id=company_id)

    if request.method == 'POST':
        # Preenche o formulário com os dados da empresa atual
        form = CompanyForm(request.POST, instance=company)
        if form.is_valid():
            # Salva as alterações no banco de dados
            form.save()
            return redirect('company_list')  # Redireciona de volta para a lista de empresas após a edição

    else:
        # Cria um novo formulário preenchido com os dados da empresa atual
        form = CompanyForm(instance=company)

    return render(request, 'company/company_edit.html', {'form': form, 'company': company})

def company_delete(request, company_id):
    # Obtém a empresa com base no ID ou retorna um erro 404 se não existir
    company = get_object_or_404(Company, id=company_id)

    if request.method == 'POST':
        # Verifica se o formulário foi enviado por POST (confirmação de exclusão)
        company.delete()  # Exclui a empresa do banco de dados
        return redirect('company_list')  # Redireciona para a lista de empresas após a exclusão

    # Se o formulário não foi enviado por POST, renderiza a página de confirmação de exclusão
    return render(request, 'company/company_delete.html', {'company': company})


def add_deliverable_view(request, company_id):
    company = Company.objects.get(pk=company_id)

    if request.method == 'POST':
        form = DeliverableForm(request.POST)
        if form.is_valid():
            deliverable = form.save(commit=False)
            deliverable.company = company
            deliverable.save()
            return redirect('company_detail', company_id=company_id)
    else:
        form = DeliverableForm()

    return render(request, 'calendar/add_deliverable.html', {'form': form, 'company': company})

def delivery_calendar_create(request):
    if request.method == 'POST':
        form = DeliveryCalendarForm(request.POST)
        if form.is_valid():
            print(form.cleaned_data)  # Print form data to debug
            form.save()
            return redirect('delivery_calendar_list')
        else:
            print(form.errors)  # Print form errors to debug
    else:
        form = DeliveryCalendarForm()
    
    return render(request, 'calendar/delivery_calendar_form.html', {'form': form})

def delivery_calendar_detail(request, pk):
    calendar = get_object_or_404(DeliveryCalendar, pk=pk)
    return render(request, 'calendar/delivery_calendar_detail.html', {'calendar': calendar})

def delivery_calendar_update(request, pk):
    calendar = get_object_or_404(DeliveryCalendar, pk=pk)
    if request.method == 'POST':
        form = DeliveryCalendarForm(request.POST, instance=calendar)
        if form.is_valid():
            form.save()
            return redirect('delivery_calendar_detail', pk=calendar.pk)
    else:
        form = DeliveryCalendarForm(instance=calendar)
    return render(request, 'calendar/delivery_calendar_form.html', {'form': form, 'calendar': calendar})

def delivery_calendar_delete(request, pk):
    calendar = get_object_or_404(DeliveryCalendar, pk=pk)
    if request.method == 'POST':
        calendar.delete()
        return redirect('delivery_calendar_list')
    return render(request, 'calendar/delivery_calendar_confirm_delete.html', {'calendar': calendar})

def delivery_calendar_list(request):
    calendars = DeliveryCalendar.objects.all()
    return render(request, 'calendar/delivery_calendar_list.html', {'calendars': calendars})

@login_required
def create_work_order(request):
    if request.method == 'POST':
        form = WorkorderForm(request.POST)
        if form.is_valid():
            work_order = form.save(commit=False)
            work_order.created_by = request.user

            if work_order.type == 'RSA billing':
                work_order.billing = 'No'

            work_order.created_at = make_aware(datetime.now())

            work_order.sla_date = sla_calculation(work_order.created_at, work_order.priority)

            work_order.save()
            return redirect('analyze_work_order')
        else:
            print("Form is not valid:", form.errors)
    else:
        form = WorkorderForm()

    companies = Company.objects.all()
    users = CustomUser.objects.all()

    context = {
        'form': form,
        'companies': companies,
        'users': users,
        'department': 'Operation',
        'type': 'Work Order',

    }

    return render(request, 'workorder/create_workorder.html', context)

@login_required
def details_workorder(request, workorder_id):
    workorder = get_object_or_404(Workorder, id=workorder_id)

    context = {
        'workorder': workorder,
        'status_choices': Workorder.STATUS_CHOICES,
    }

    return render(request, 'workorder/details_workorder.html', context)

import logging
@login_required
def analyze_work_order(request):
    if request.method == 'POST':
        workorder_id = request.POST.get('workorder_id')
        cam_approval = request.POST.get('cam_approval')
        billing = request.POST.get('billing')
        workorder = Workorder.objects.get(id=workorder_id)

        logger = logging.getLogger(__name__)
        logger.info('Saving work order %d', workorder.id)

        if request.user.is_cam:
            if cam_approval == 'Approve':
                workorder.status = 'In Progress'
                workorder.cam_approval = 'Approved'
                workorder.department = 'Operation'
                workorder.billing = billing
                workorder.approve_date = now()
                workorder.save()
                return redirect('execute_work_order')
            elif cam_approval == 'Reprove':
                workorder.status = 'Rework'
                workorder.cam_approval = 'Reproved'
                workorder.assigned_to = workorder.created_by
                workorder.billing = billing
                workorder.approve_date = None
                workorder.save()
                return redirect('rework_work_order')
        else:
            return redirect('home')

    workorders = Workorder.objects.filter(status='New')

    context = {
        'workorders': workorders,
        'users': CustomUser.objects.all(),
    }

    return render(request, 'workorder/analyze_workorder.html', context)

#nessa função é para criar um ticket automaticamente quando a WO for aprovada pelo CAM e vai para Execute
@receiver(post_save, sender=Workorder)
def create_tickets(sender, instance, **kwargs):
    if instance.cam_approval == 'Approved':
        Ticket.objects.create(
            title=f'WO - Execute #{instance.id}',
            description=f'Hi, this is an automatic ticket to execute WO approved.',
            priority='Medium',
            assigned_to=instance.assigned_to,
            created_by=instance.created_by,
            company=instance.company,
            department='Operation',
            type='Work Order',
            sla_date=instance.sla_date,
        )

@login_required
def execute_work_order(request):
    if request.method == 'POST':
        workorder_id = request.POST.get('workorder_id')
        workorder = Workorder.objects.get(id=workorder_id)
        workorder.status = 'Executed'
        workorder.complete_date = now()
        workorder.save()
        messages.success(request, 'Work Order has been executed.')
        return redirect('work_order_dashboard')
    
    workorders = Workorder.objects.filter(status='In Progress')

    context = {
        'workorders': workorders,
    }

    return render(request, 'workorder/execute_workorder.html', context)

@login_required
def rework_work_order(request):
    if request.method == 'POST':
        workorder_id = request.POST.get('workorder_id')
        workorder = Workorder.objects.get(id=workorder_id)
        workorder.status = 'New'
        workorder.cam_approval = 'Reproved'
        workorder.assigned_to = request.user
        workorder.save()
        messages.success(request, 'Work Order has been sent for rework.')
        return redirect('analyze_work_order')

    workorders = Workorder.objects.filter(status='Rework')

    context = {
        'workorders': workorders,
        'users': CustomUser.objects.all(),
    }

    return render(request, 'workorder/rework_workorder.html', context)

@login_required
def work_order_dashboard(request):
    workorders = Workorder.objects.all()

    context = {
        'workorders': workorders,
    }

    return render(request, 'workorder/workorder_dashboard.html', context)

@login_required
def edit_work_order(request, workorder_id):
    workorder = get_object_or_404(Workorder, id=workorder_id)

    if request.method == 'POST':
        form = WorkorderForm(request.POST, instance=workorder)
        if form.is_valid():
            form.save()
            messages.success(request, 'Work Order has been updated.')
            return redirect('rework_work_order')
    else:
        form = WorkorderForm(instance=workorder)

    context = {
        'form': form,
    }

    return render(request, 'workorder/edit_workorder.html', context)

@login_required
def report_workorder(request):
    # Obtenha todos os analistas
    analysts = CustomUser.objects.filter(is_analyst=True).distinct()

    # Inicialize uma lista para armazenar os dados dos analistas com horas faturadas
    analyst_data = []

    # Obtenha os meses distintos das Work Orders
    months = Workorder.objects.annotate(month=ExtractMonth('created_at')).values('month').distinct()

    for analyst in analysts:
        # Filtrar as Work Orders do analista para cada mês
        analyst_workorders = Workorder.objects.filter(created_by=analyst, status='Executed')
        
        analyst_month_data = {'analyst': analyst, 'months': []}
        
        for month in months:
            # Filtrar as Work Orders do analista para o mês atual
            month_workorders = analyst_workorders.filter(created_at__month=month['month'])
            
            total_hours = month_workorders.aggregate(Sum('hours'))['hours__sum']
            billing_count = month_workorders.filter(billing='Yes').count()
            
            # Verificar se o analista teve horas faturadas para este mês
            if billing_count > 0:
                analyst_month_data['months'].append({
                    'month': month['month'],
                    'total_hours': total_hours or 0,
                    'billing_count': billing_count,
                    'billed': True,
                })

        # Se o analista teve horas faturadas em algum mês, adicione-o à lista
        if analyst_month_data['months']:
            analyst_data.append(analyst_month_data)

    context = {
        'analyst_data': analyst_data,
    }
    return render(request, 'workorder/report_workorder.html', context)



# Função para validar arquivo XML de NFe
def validate_xml_nfe(xml_content):
    try:
        # Adicione sua lógica de validação XML aqui
        # Por exemplo, você pode usar a biblioteca lxml para fazer a validação XML.
        # Substitua este exemplo pela sua lógica de validação específica.
        from lxml import etree
        xml_parser = etree.XMLParser(dtd_validation=True)
        etree.fromstring(xml_content, xml_parser)
        return True
    except Exception as e:
        return False

# Função para validar arquivo PDF de NFe
def validate_pdf_nfe(pdf_content):
    try:
        # Verifique se o arquivo tem extensão PDF
        if not pdf_content.name.endswith('.pdf'):
            return False, "O arquivo não tem a extensão PDF."

        pdf_data = pdf_content.read()
        if b'nf' not in pdf_data:
            return False, "O arquivo PDF não contém a palavra 'NF-e'."

        chave_acesso = capture_chave_acesso(pdf_data)

        if not re.match(r'^\d{44}$', chave_acesso):
            return False, "A chave de acesso não possui o formato correto."

        pdf_file_path = os.path.join(settings.MEDIA_ROOT, 'pdf_files', f'{chave_acesso}.pdf')
        with open(pdf_file_path, 'wb') as destination:
            destination.write(pdf_data)

        return True, "Arquivo PDF de NFe válido e salvo com sucesso."
    except Exception as e:
        return False, f"Erro ao validar o arquivo PDF: {str(e)}"

# Função para capturar a chave de acesso do PDF (substitua pela sua lógica)
def capture_chave_acesso(pdf_content):

    # Substitua esta lógica pela forma como você extrairá a chave de acesso do PDF.
    # Aqui estamos apenas fazendo uma simulação.
    chave_acesso = '12345678901234567890123456789012345678901234'
    return chave_acesso

# View para fazer upload de arquivos XML e PDF
def read_nfe(request):
    if request.method == 'POST':
        # Verifique se um arquivo XML foi enviado
        if 'xml_file' in request.FILES:
            xml_file = request.FILES['xml_file']

        if xml_file:
            # Chame a função para analisar o XML
            invoice = parse_invoice_xml(xml_file)

            if invoice:
                # O XML foi analisado com sucesso e uma fatura foi criada
                return redirect('invoice_list')
            else:
                error_message = "Ocorreu um erro ao analisar o XML."

            return render(request, 'nfes/read_nfe.html', {'error_message': error_message})

        # Verifique se um arquivo PDF foi enviado
        if 'pdf_file' in request.FILES:
            pdf_file = request.FILES['pdf_file']

            # Valide o arquivo PDF de NFe
            is_valid, message = validate_pdf_nfe(pdf_file)

            if is_valid:
                return JsonResponse({'message': message})
            else:
                return JsonResponse({'error': message}, status=400)

        return JsonResponse({'message': 'Arquivos enviados com sucesso.'})

    return render(request, 'nfes/read_nfe.html')

def parse_invoice_xml(xml_file):
    try:
        ns = {"nfe": "http://www.portalfiscal.inf.br/nfe"}
        tree = ET.parse(xml_file)
        root = tree.getroot()
        number = root.find(".//nfe:ide/nfe:nNF", namespaces=ns).text
        series = root.find(".//nfe:ide/nfe:serie", namespaces=ns).text
        issuance_date_str = root.find(".//nfe:ide/nfe:dhEmi", namespaces=ns).text
        issuance_date = datetime.fromisoformat(issuance_date_str)
        exit_date_str = root.find(".//nfe:ide/nfe:dhSaiEnt", namespaces=ns).text
        exit_date = datetime.fromisoformat(exit_date_str)
        emitter_cnpj = root.find(".//nfe:emit/nfe:CNPJ", namespaces=ns).text
        emitter_name = root.find(".//nfe:emit/nfe:xNome", namespaces=ns).text
        recipient_cnpj = root.find(".//nfe:dest/nfe:CNPJ", namespaces=ns).text
        recipient_name = root.find(".//nfe:dest/nfe:xNome", namespaces=ns).text
        total_value = root.find(".//nfe:total/nfe:ICMSTot/nfe:vNF", namespaces=ns).text

        # Verifique se o motorista (emitente) já existe no sistema com base no CPF ou CNPJ
        driver = CustomUser.objects.filter(Q(cpf=emitter_cnpj) | Q(cnpj=emitter_cnpj)).first()

        if driver:
            # Emitente cadastrado, atualize a coluna 'driver' da fatura com o motorista existente
            invoice = Invoice.objects.create(
                number=number,
                series=series,
                issuance_date=issuance_date,
                exit_date=exit_date,
                emitter_cnpj=emitter_cnpj,
                emitter_name=emitter_name,
                recipient_cnpj=recipient_cnpj,
                recipient_name=recipient_name,
                total_value=total_value,
                status='Finance Approval',  # Defina o status conforme necessário
                driver=driver
            )
        else:
            # Emitente não cadastrado, crie uma nova fatura com status 'Driver Check'
            invoice = Invoice.objects.create(
                number=number,
                series=series,
                issuance_date=issuance_date,
                exit_date=exit_date,
                emitter_cnpj=emitter_cnpj,
                emitter_name=emitter_name,
                recipient_cnpj=recipient_cnpj,
                recipient_name=recipient_name,
                total_value=total_value,
                status='Driver Check'  # Defina o status conforme necessário
            )

        return invoice

    except Exception as e:
        print(f"Ocorreu um erro ao analisar o XML: {str(e)}")
        return None

def create_manual_invoice(request):
    if request.method == 'POST':
        form = ManualInvoiceForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('invoice_list')
    else:
        form = ManualInvoiceForm()
    
    return render(request, 'nfes/manual_invoice.html', {'form': form})

def invoice_list(request):
    invoices = Invoice.objects.all()
    return render(request, 'nfes/invoice_list.html', {'invoices': invoices})

def invoice_edit(request, pk):
    invoice = get_object_or_404(Invoice, pk=pk)

    if request.method == 'POST':
        form = ManualInvoiceForm(request.POST, instance=invoice)
        if form.is_valid():
            form.save()
            return redirect('invoice_list')  # Redirecionar após a edição bem-sucedida
    else:
        form = ManualInvoiceForm(instance=invoice)

    return render(request, 'nfes/invoice_edit.html', {'form': form})

def invoice_delete(request, pk):
    invoice = get_object_or_404(Invoice, pk=pk)

    if request.method == 'POST':
        invoice.delete()
        return redirect('invoice_list')  # Redirecionar após a exclusão bem-sucedida

    return render(request, 'nfes/invoice_delete.html', {'invoice': invoice})

def driver_check(request):
    if request.method == 'POST':
        invoice_id = request.POST.get('invoice_id')
        operation_approval = request.POST.get('operation_approval')
        invoice = get_object_or_404(Invoice, id=invoice_id)
        
        if operation_approval == 'Approve':
            invoice.status = 'Finance Approval'
            invoice.operation_approval = 'Approved'
            invoice.approve_date = now()  # Define a data de aprovação
            invoice.save()
            return redirect('invoice_list')  # Redirecione para a lista de invoices após a aprovação
        elif operation_approval == 'Reject':
            invoice.status = 'Driver Check'
            invoice.operation_approval = 'Rejected'
            invoice.approve_date = None  # Define a data de aprovação como nula ao rejeitar
            invoice.save()
            return redirect('invoice_list')  # Redirecione para a lista de invoices após a rejeição

    invoices = Invoice.objects.filter(status='Driver Check')

    context = {
        'invoices': invoices,
    }

    return render(request, 'nfes/invoice_driver_check.html', context)

def invoice_finance_approval(request):
    if request.method == 'POST':
        invoice_id = request.POST.get('invoice_id')
        finance_approval = request.POST.get('finance_approval')
        invoice = get_object_or_404(Invoice, id=invoice_id)
        
        if finance_approval == 'Approve':
            invoice.status = 'Approved'
            invoice.finance_approval = 'Approved'
            invoice.approve_date = now()  # Define a data de aprovação
            invoice.save()
            return redirect('invoice_list')  # Redirecione para a lista de invoices após a aprovação
        elif finance_approval == 'Reject':
            invoice.status = 'Rejected'
            invoice.finance_approval = 'Rejected'
            invoice.approve_date = None  # Define a data de aprovação como nula ao rejeitar
            invoice.save()
            return redirect('invoice_list')  # Redirecione para a lista de invoices após a rejeição

    invoices = Invoice.objects.filter(status='Finance Approval')

    context = {
        'invoices': invoices,
    }

    return render(request, 'nfes/invoice_finance_approval.html', context)

def create_bank(request):
    if request.method == 'POST':
        form = BankForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('home')  # Redireciona para a lista de bancos ou outra página desejada após o cadastro
    else:
        form = BankForm()
    
    return render(request, 'finance/create_bank.html', {'form': form})

def changes_upload(request):
    if request.method == 'POST':
        form = ChangesUploadForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = request.FILES['excel_file']
            company = form.cleaned_data['company'].id
            result_message = changes_data(excel_file, company)

            return render(request, 'validation/changes_upload.html', {'form': form, 'result_message': result_message})
    else:
        form = ChangesUploadForm()
    
    return render(request, 'validation/changes_upload.html', {'form': form})

def extract_date_from_pay_period(pay_period):
    match = re.search(r'\d{4}-\d{2}-\d{2}', pay_period)
    if match:
        return match.group()
    return None

def convert_to_date(date_str):
    try:
        if not pd.isna(date_str):
            return datetime.strptime(str(date_str), '%Y-%m-%d').date()
    except ValueError:
        pass

    return None

def convert_to_decimal(decimal_str):
    try:
        if not pd.isna(decimal_str):
            return float(decimal_str)
    except ValueError:
        pass

    return None

def convert_to_int(integer_str):
    try:
        if not pd.isna(integer_str):
            return int(integer_str)
    except ValueError:
        pass
    return None

def changes_data(file_path, company_id):
    try:
        df = pd.read_excel(file_path)

        # Nomes das colunas no arquivo Excel correspondentes aos nomes do modelo
        column_mapping = {
            'Payroll Name': 'Payroll Name',
            'Employee Name': 'Employee Name',
            'Employee ID Number': 'Employee ID Number',
            'Pay Period': 'Pay Period',
            'Payroll Pay Element': 'Payroll Pay Element',
            'Transaction Type': 'Transaction Type',
            'Element Type': 'Element Type',
            'Effective Date': 'Effective Date',
            'End Date': 'End Date',
            'Amount': 'Amount',
            'Number': 'Number',
            'Unit Code': 'Unit Code',
            'Unit': 'Unit',
            'Rate': 'Rate',
            'Payroll Effective Date': 'Payroll Effective Date',
            'Termination Date': 'Termination Date',
            'Comments': 'Comments',
            'ICP EEPR ID': 'ICP EEPR ID',
            'ICP Payroll PayElement': 'ICP Payroll PayElement',
            'ICP Pay Element Code': 'ICP Pay Element Code',
        }

        # Renomeie as colunas do DataFrame para corresponder ao modelo
        df.rename(columns=column_mapping, inplace=True)

        required_columns = list(column_mapping.values())
        missing_columns = [col for col in required_columns if col not in df.columns]

        if missing_columns:
            return f"Colunas ausentes: {', '.join(missing_columns)}"

        # Obtém o objeto da empresa com base no ID
        company = Company.objects.get(id=company_id)

        for _, row in df.iterrows():
            pay_period = extract_date_from_pay_period(row['Pay Period'])
            effective_date = convert_to_date(row['Effective Date'])
            end_date = convert_to_date(row['End Date'])
            payroll_effective_date = convert_to_date(row['Payroll Effective Date'])
            termination_date = convert_to_date(row['Termination Date'])
            amount = convert_to_decimal(row['Amount'])
            number = convert_to_decimal(row['Number'])
            unit = convert_to_decimal(row['Unit'])
            rate = convert_to_decimal(row['Rate'])
            icp_eepr_id = row['ICP EEPR ID'] if not pd.isna(row['ICP EEPR ID']) else 0
            icp_payroll_payelement = convert_to_int(row['ICP Payroll PayElement'])

            ChangesValidation.objects.create(
                Payroll_Name=row['Payroll Name'],
                Employee_Name=row['Employee Name'],
                Employee_ID_Number=row['Employee ID Number'],
                Pay_Period=pay_period,
                Payroll_Pay_Element=row['Payroll Pay Element'],
                Transaction_Type=row['Transaction Type'],
                Element_Type=row['Element Type'],
                Effective_Date=effective_date,
                End_Date=end_date,
                Amount=amount,
                Number=number,
                Unit_Code=row['Unit Code'],
                Unit=unit,
                Rate=rate,
                Payroll_Effective_Date=payroll_effective_date,
                Termination_Date=termination_date,
                Comments=row['Comments'],
                ICP_EEPR_ID=icp_eepr_id,
                ICP_Payroll_PayElement=icp_payroll_payelement,
                ICP_Pay_Element_Code=row['ICP Pay Element Code'],
                company=company
            )

        return f"Dados importados com sucesso! Registros importados: {len(df)} (Payroll Name: {', '.join(df['Payroll Name'].unique())}, Pay Period: {', '.join(df['Pay Period'].unique())})"

    except Exception as e:
        return f"Erro ao importar os dados: {str(e)}"

def delete_changes(request):
    message = None
    companies = Company.objects.all()
    payrolls = []
    pay_periods = []

    if request.method == 'POST':
        company_id = request.POST.get('company')
        payroll_name = request.POST.get('payroll_name')
        pay_period = request.POST.get('pay_period')

        # Exclua registros com base na empresa, Payroll Name e Pay Period
        deleted_count = ChangesValidation.objects.filter(
            company_id=company_id,
            Payroll_Name=payroll_name,
            Pay_Period=pay_period
        ).delete()[0]

        # Mensagem sobre os registros que foram apagados e a quantidade apagada
        message = f"{deleted_count} registros foram apagados."

        # Recarregue a lista de payrolls e pay_periods após a exclusão
        payrolls = ChangesValidation.objects.filter(company_id=company_id).values_list('Payroll_Name', flat=True).distinct()
        pay_periods = ChangesValidation.objects.filter(company_id=company_id).values_list('Pay_Period', flat=True).distinct()

    return render(request, 'validation/delete_changes.html', {
        'companies': companies,
        'message': message,
        'payrolls': payrolls,
        'pay_periods': pay_periods,
    })

def upload_generate(request):
    if request.method == 'POST':
        company_id = request.POST.get('company')
        pay_period = request.POST.get('pay_period')
        generate_type = request.POST.get('generate_type')

        if generate_type == 'single_variables':
            return generate_single_variables(request, company_id, pay_period)
        elif generate_type == 'single_overtime':
            return generate_single_overtime(request, company_id, pay_period)
        elif generate_type == 'recurring':
            return generate_recurring(request, company_id, pay_period)

    companies = Company.objects.all()
    return render(request, 'validation/upload_generate.html', {'companies': companies})

def generate_single_variables(request, company_id, pay_period):
    if request.method == 'POST':
        company = get_object_or_404(Company, id=company_id)
        company_abbr = company.abreviation

        single_variables_records = ChangesValidation.objects.filter(
            company_id=company_id,
            Pay_Period=pay_period,
            Element_Type='Single',
            Amount__isnull=False
        ).values('ICP_EEPR_ID', 'Amount', 'ICP_Payroll_PayElement', 'Employee_ID_Number')

        filename = f"{company_abbr}_Bulk_Single_{pay_period}.csv"

        # Criar a resposta HTTP para o download da planilha CSV
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        # Criar o escritor CSV
        writer = csv.writer(response, delimiter=';')

        # Escrever o cabeçalho
        writer.writerow(['ICP EEPR ID', 'ICP Payroll PayElement', 'Amount'])

        for record in single_variables_records:
            # Verificar se ICP_EEPR_ID está vazio ou igual a 0 e, em seguida, use Employee_ID_Number
            icp_eepr_id = record['ICP_EEPR_ID'] if record['ICP_EEPR_ID'] is not None and record['ICP_EEPR_ID'] != 0 else record['Employee_ID_Number']

            # Converter Amount substituindo . por , se não for None
            amount = str(record['Amount']).replace('.', ',') if record['Amount'] is not None else ''

            writer.writerow([icp_eepr_id, record['ICP_Payroll_PayElement'], amount])

        return response

    companies = Company.objects.all()
    return render(request, 'validation/upload_generate.html', {'companies': companies})

def generate_single_overtime(request, company_id, pay_period):
    if request.method == 'POST':
        company = get_object_or_404(Company, id=company_id)
        company_abbr = company.abreviation

        single_overtime_records = ChangesValidation.objects.filter(
            company_id=company_id,
            Pay_Period=pay_period,
            Element_Type='Single',
            Number__isnull=False
        ).values('ICP_EEPR_ID', 'Number', 'ICP_Payroll_PayElement', 'Employee_ID_Number')

        filename = f"{company_abbr}_Bulk_Single_Overtime_{pay_period}.csv"

        # Criar a resposta HTTP para o download da planilha CSV
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        writer = csv.writer(response, delimiter=';')
        writer.writerow(['ICP EEPR ID', 'ICP Payroll PayElement', 'Number'])

        for record in single_overtime_records:
            icp_eepr_id = record['ICP_EEPR_ID'] if record['ICP_EEPR_ID'] is not None and record['ICP_EEPR_ID'] != 0 else record['Employee_ID_Number']
            
            number = str(record['Number']).replace('.', ',') if record['Number'] is not None else ''
            
            writer.writerow([icp_eepr_id, record['ICP_Payroll_PayElement'], number])

        return response

    companies = Company.objects.all()
    return render(request, 'validation/upload_generate.html', {'companies': companies})

def generate_recurring(request, company_id, pay_period):
    if request.method == 'POST':
        company = get_object_or_404(Company, id=company_id)
        company_abbr = company.abreviation

        recurring_records = ChangesValidation.objects.filter(
            company_id=company_id,
            Pay_Period=pay_period,
            Element_Type='Recurring',
            Amount__isnull=False
        ).values('ICP_EEPR_ID', 'Amount', 'ICP_Payroll_PayElement', 'Employee_ID_Number')

        filename = f"{company_abbr}_Bulk_Recurring_{pay_period}.csv"

        # Criar a resposta HTTP para o download da planilha CSV
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        # Criar o escritor CSV
        writer = csv.writer(response, delimiter=';')

        # Escrever o cabeçalho
        writer.writerow(['ICP EEPR ID', 'ICP Payroll PayElement', 'Amount'])

        for record in recurring_records:
            # Verificar se ICP_EEPR_ID está vazio ou igual a 0 e, em seguida, use Employee_ID_Number
            icp_eepr_id = record['ICP_EEPR_ID'] if record['ICP_EEPR_ID'] is not None and record['ICP_EEPR_ID'] != 0 else record['Employee_ID_Number']

            # Converter Amount substituindo . por , se não for None
            amount = str(record['Amount']).replace('.', ',') if record['Amount'] is not None else ''

            writer.writerow([icp_eepr_id, record['ICP_Payroll_PayElement'], amount])

        return response

    companies = Company.objects.all()
    return render(request, 'validation/upload_generate.html', {'companies': companies})

def write_csv(response, header, data):
    writer = csv.writer(response, delimiter=';')
    writer.writerow(header)
    for row in data:
        writer.writerow(row)


def pay_elements_upload(request):
    companies = Company.objects.all() 
    
    if request.method == 'POST':
        form = ChangesUploadForm(request.POST, request.FILES)
        if form.is_valid():
            csv_file = request.FILES['csv_file']
            company = form.cleaned_data['company']
            result_message = pay_elements_data(csv_file, company)

            return render(request, 'validation/pay_elements_upload.html', {'form': form, 'result_message': result_message})
    else:

        form = ChangesUploadForm()

    return render(request, 'validation/pay_elements_upload.html', {'form': form, 'companies': companies})
def pay_elements_data(file_path, company_id):
    try:
        df = pd.read_csv(file_path)

        column_mapping = {
            'Payroll Name': 'Payroll Name',
            'Pay Period': 'Pay Period',
            'Payroll Pay Element': 'Payroll Pay Element',
            'Employee ID': 'Employee ID',
            'Amount': 'Amount',
            'Number': 'Number',
            'Unit Code': 'Unit Code',
            'Unit': 'Unit',
            'Rate': 'Rate',
            'ICP Pay Element Code': 'ICP Pay Element Code',
            'Local ID': 'Local ID',
            'ICP Payroll Name': 'ICP Payroll Name',
        }

        df.rename(columns=column_mapping, inplace=True)

        required_columns = list(column_mapping.values())
        missing_columns = [col for col in required_columns if col not in df.columns]

        if missing_columns:
            return f"Colunas ausentes: {', '.join(missing_columns)}"

        company = Company.objects.get(id=company_id)

        for _, row in df.iterrows():
            pay_period = row['Pay Period']
            amount = row['Amount']
            number = row['Number']
            unit = row['Unit']
            rate = row['Rate']
            ICP_Pay_Element_Code = row['ICP Pay Element Code']

            ChangesValidation.objects.create(
                Payroll_Name=row['Payroll Name'],
                Employee_ID=row['Employee ID'],
                Pay_Period=pay_period,
                Amount=amount,
                Number=number,
                Unit_Code=row['Unit Code'],
                Unit=unit,
                Rate=rate,
                Local_ID=row['Local ID'],
                ICP_Pay_Element_Code=ICP_Pay_Element_Code,
                company=company,
                ICP_Payroll_Name=row['ICP Payroll Name'],
            )

        return f"Dados importados com sucesso! Registros importados: {len(df)} (Payroll Name: {', '.join(df['Payroll Name'].unique())}, Pay Period: {', '.join(df['Pay Period'].unique())})"

    except Exception as e:
        return f"Erro ao importar os dados: {str(e)}"
